#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sqlite3
import sys
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCHEMA_VERSION = 1
MINIMUM_APP_VERSION = None
SQLITE_NAME = "reimbursement.sqlite"
ZIP_NAME = "reimbursement.zip"
MANIFEST_NAME = "reimbursement_manifest.json"
PIPELINE_ID = "docspace-reimbursement-manual-xlsx-v1"
CANONICAL_SOURCE_ID = "manual-input:reimbursement-xlsx"
MINIMUM_RELATIVE_COUNT_RATIO = 0.80
# Curated ICD-10 mappings are hand-made and cannot be regenerated from the workbook,
# so the transfer is held to a much tighter bound than row volume. The ratio is measured
# against *eligible* codes — previous rows whose stable key is unique in the previous
# database — because rows with a duplicated key are structurally untransferable and must
# not be charged against a genuine key-matching regression.
MINIMUM_ICD10_RETENTION_RATIO = 0.90
VOLUME_CHECK_SECTIONS = ("medicines", "insulins", "combined_medicines")
# Tables this pipeline rewrites from the workbook. Everything else in the baseline is
# carried over untouched and must survive the build.
OWNED_TABLES = {"reimbursement_items", "metadata"}

UA_MONTHS = {
    "січня": 1,
    "лютого": 2,
    "березня": 3,
    "квітня": 4,
    "травня": 5,
    "червня": 6,
    "липня": 7,
    "серпня": 8,
    "вересня": 9,
    "жовтня": 10,
    "листопада": 11,
    "грудня": 12,
}

SECTIONS = {
    "medicines": "I. Лікарські засоби, крім препаратів інсуліну та комбінованих лікарських засобів",
    "insulins": "II. Препарати інсуліну",
    "combined_medicines": "III. Комбіновані лікарські засоби",
}

# SQL type per item column, so a column added to an older baseline keeps REAL affinity
# instead of degrading numbers to TEXT (which breaks ORDER BY and range queries).
ITEM_COLUMN_TYPES = {
    "source_row": "INTEGER",
    "section": "TEXT",
    "section_uk": "TEXT",
    "ordinal_no": "INTEGER",
    "inn_name": "TEXT",
    "trade_name": "TEXT",
    "dosage_form": "TEXT",
    "dosage": "TEXT",
    "pack_units": "REAL",
    "atc_code": "TEXT",
    "manufacturer_country": "TEXT",
    "registration_number": "TEXT",
    "registration_expiry": "TEXT",
    "wholesale_price_uah": "REAL",
    "retail_price_uah": "REAL",
    "who_daily_dose": "REAL",
    "reimbursement_daily_dose_uah": "REAL",
    "reimbursement_primary_pack_uah": "REAL",
    "reimbursement_release_form_or_primary_pack_uah": "REAL",
    "reimbursement_pack_uah": "REAL",
    "patient_copay_uah": "REAL",
    "copay_type": "TEXT",
    "release_form_or_primary_pack": "TEXT",
    "icd10_codes": "TEXT",
}

ITEM_COLUMNS = [
    "source_row",
    "section",
    "section_uk",
    "ordinal_no",
    "inn_name",
    "trade_name",
    "dosage_form",
    "dosage",
    "pack_units",
    "atc_code",
    "manufacturer_country",
    "registration_number",
    "registration_expiry",
    "wholesale_price_uah",
    "retail_price_uah",
    "who_daily_dose",
    "reimbursement_daily_dose_uah",
    "reimbursement_primary_pack_uah",
    "reimbursement_release_form_or_primary_pack_uah",
    "reimbursement_pack_uah",
    "patient_copay_uah",
    "copay_type",
    "release_form_or_primary_pack",
    "icd10_codes",
]

STABLE_KEY_FIELDS = (
    "section",
    "registration_number",
    "trade_name",
    "dosage_form",
    "dosage",
    "pack_units",
    "atc_code",
)

REQUIRED_ITEM_COLUMNS = {
    "id",
    "source_row",
    "section",
    "section_uk",
    "inn_name",
    "trade_name",
    "dosage_form",
    "dosage",
    "pack_units",
    "atc_code",
    "manufacturer_country",
    "registration_number",
    "registration_expiry",
    "retail_price_uah",
    "reimbursement_pack_uah",
    "patient_copay_uah",
    "icd10_codes",
}

FREE_COPAY_LABELS = {"безоплатно"}


@dataclass(frozen=True)
class SourceCandidate:
    name: str
    url: str
    resource_id: str
    register_date: date
    description: str


@dataclass(frozen=True)
class BuildStats:
    source_version: str
    source_file: str
    rows_processed: int
    medicines_count: int
    insulins_count: int
    combined_medicines_count: int
    medical_devices_count: int
    database_version: int
    sha256: str
    icd10_codes_transferred: int
    icd10_codes_previous: int
    icd10_codes_eligible: int
    baseline: str
    expected_release_tag: str
    release_created: bool


def log(message: str) -> None:
    print(message, flush=True)


def parse_date_from_text(text: str) -> date | None:
    compact = re.search(r"(\d{1,2})[._-](\d{1,2})[._-](20\d{2})", text)
    if compact:
        day, month, year = map(int, compact.groups())
        return date(year, month, day)

    spoken = re.search(
        r"(\d{1,2})\s+("
        + "|".join(UA_MONTHS)
        + r")\s+(20\d{2})(?:\s+року)?",
        text.casefold(),
    )
    if spoken:
        day = int(spoken.group(1))
        month = UA_MONTHS[spoken.group(2)]
        year = int(spoken.group(3))
        return date(year, month, day)
    return None


def parse_strict_register_date_from_text(text: str) -> date | None:
    match = re.search(r"станом\s+на\s+(.{0,60})", text.casefold())
    if not match:
        return None
    return parse_date_from_text(match.group(1))


def parse_manifest_date(value: str | None) -> date | None:
    if not value:
        return None
    for candidate in (value[:10], value):
        try:
            return datetime.fromisoformat(candidate).date()
        except ValueError:
            pass
    return None


def latest_published_manifest(path: Path | None) -> dict[str, Any] | None:
    if path is None or not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def find_input_workbook(input_dir: Path) -> Path:
    if not input_dir.is_dir():
        raise RuntimeError(f"Manual reimbursement input directory is missing: {input_dir}")
    candidates = sorted(
        path for path in input_dir.iterdir() if path.is_file() and path.suffix.casefold() == ".xlsx"
    )
    if len(candidates) != 1:
        raise RuntimeError(
            "Manual reimbursement input must contain exactly one XLSX file: "
            f"found {len(candidates)} in {input_dir}"
        )
    return candidates[0]


def normalized_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ").strip())


def normalized_cell(value: Any) -> str:
    return normalized_text(value).casefold().replace("’", "'")


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return normalized_text(value)


def as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).replace(",", ".")))
    except ValueError:
        return None


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        if not math.isfinite(number):
            raise RuntimeError(f"Unexpected non-finite numeric value: {value!r}")
        return number

    text = normalized_text(value).replace("\u00a0", " ")
    # A space between digit groups is the Ukrainian thousands separator and carries no
    # ambiguity, so it is dropped before anything else.
    compact = re.sub(r"(?<=\d)\s+(?=\d)", "", text)
    compact = re.sub(r"\s+", "", compact)

    if "," in compact and "." in compact:
        # Both separators present: the right-most one is the decimal separator.
        normalized = (
            compact.replace(".", "").replace(",", ".")
            if compact.rfind(",") > compact.rfind(".")
            else compact.replace(",", "")
        )
    else:
        separator = "," if "," in compact else ("." if "." in compact else "")
        if not separator:
            normalized = compact
        elif compact.count(separator) > 1:
            # 1,234,567 / 1.234.567 — repeated separators can only be grouping.
            normalized = compact.replace(separator, "")
        elif re.fullmatch(r"[+-]?\d{1,3}" + re.escape(separator) + r"\d{3}", compact):
            # 1,234 is either 1234 (grouping) or 1.234 (decimal) and nothing in the cell
            # tells us which. Guessing silently corrupts prices, so refuse.
            raise RuntimeError(
                f"Ambiguous numeric value {value!r}: {compact!r} could be either a decimal "
                "separator or a thousands separator. Fix the source cell formatting."
            )
        else:
            normalized = compact.replace(separator, ".")

    if not re.fullmatch(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)", normalized):
        raise RuntimeError(f"Unexpected numeric value: {value!r}")
    return float(normalized)


def parse_patient_copay(value: Any) -> tuple[float | None, str]:
    if normalized_cell(value) in FREE_COPAY_LABELS:
        return 0.0, normalized_text(value)
    return as_float(value), ""


def normalized_key_number(value: Any) -> str:
    number = as_float(value)
    if number is None:
        return ""
    return format(number, ".15g")


def stable_item_key(row: dict[str, Any] | sqlite3.Row) -> tuple[str, ...] | None:
    parts: list[str] = []
    for field in STABLE_KEY_FIELDS:
        value = row[field]
        if field == "pack_units":
            normalized = normalized_key_number(value)
        else:
            normalized = normalized_cell(value)
        if not normalized:
            return None
        parts.append(normalized)
    return tuple(parts)


def normalize_roman_numeral(numeral: str) -> str:
    """Cyrillic І/і are visually identical to Latin I/i and both occur in these workbooks.

    Applied to the numeral only — folding the whole heading would also rewrite Ukrainian
    words such as "Лікарські" and break the keyword check.
    """
    return numeral.replace("І", "I").replace("і", "i").casefold()


SECTION_ROMAN_NUMERALS = {"i": "medicines", "ii": "insulins", "iii": "combined_medicines"}

SECTION_HEADER_KEYWORDS = {
    "medicines": ("лікарські засоби", "крім"),
    "insulins": ("препарати інсуліну",),
    "combined_medicines": ("комбінован",),
}

SECTION_HEADER_MAX_CELLS = 3
SECTION_HEADER_RE = re.compile(r"^\s*([iIіІ]{1,3})\s*[.)]")

# Column resolution is header driven: every field names the substrings that must all appear in
# the header cell. `exact` disambiguates fields whose header is a prefix of another one
# ("Форма випуску" vs "Форма випуску/ первинна упаковка").
COMMON_COLUMN_SPECS: dict[str, dict[str, Any]] = {
    "ordinal_no": {"exact": ("№", "№ з/п", "порядковий номер")},
    "inn_name": {"all": ("міжнародна", "непатентована")},
    "trade_name": {"all": ("торговельна", "назва")},
    "dosage_form": {"all": ("форма випуску",), "exact": ("форма випуску",)},
    "dosage": {"all": ("дозування",)},
    "pack_units": {"all": ("кількість",)},
    "atc_code": {"all": ("код атх",)},
    "manufacturer_country": {"all": ("найменування виробника",)},
    "registration_number": {"all": ("номер реєстраційного",)},
    "registration_expiry": {"all": ("дата закінчення",)},
    "wholesale_price_uah": {"all": ("оптово",)},
    "retail_price_uah": {"all": ("роздрібна",)},
    "patient_copay_uah": {"all": ("сума доплати",)},
}

SECTION_COLUMN_SPECS: dict[str, dict[str, dict[str, Any]]] = {
    "medicines": {
        **COMMON_COLUMN_SPECS,
        "who_daily_dose": {"all": ("добова доза",)},
        "reimbursement_daily_dose_uah": {"all": ("розмір реімбурсації", "добової")},
        "reimbursement_pack_uah": {"all": ("розмір реімбурсації", "за упаковку")},
    },
    "insulins": {
        **COMMON_COLUMN_SPECS,
        "reimbursement_primary_pack_uah": {"all": ("розмір реімбурсації", "первинну упаковку")},
        "copay_type": {"all": ("тип доплати",)},
    },
    "combined_medicines": {
        **COMMON_COLUMN_SPECS,
        "release_form_or_primary_pack": {"all": ("форма випуску", "первинна упаковка")},
        "reimbursement_release_form_or_primary_pack_uah": {"all": ("розмір реімбурсації", "форми випуску")},
        "reimbursement_pack_uah": {"all": ("розмір реімбурсації", "за упаковку")},
    },
}

# Fields that must hold a number for the row to be accepted.
REQUIRED_NUMERIC_FIELDS = {
    "medicines": (
        "pack_units",
        "wholesale_price_uah",
        "retail_price_uah",
        "reimbursement_daily_dose_uah",
        "reimbursement_pack_uah",
        "patient_copay_uah",
    ),
    "insulins": (
        "pack_units",
        "wholesale_price_uah",
        "retail_price_uah",
        "reimbursement_primary_pack_uah",
        "patient_copay_uah",
    ),
    "combined_medicines": (
        "pack_units",
        "wholesale_price_uah",
        "retail_price_uah",
        "reimbursement_release_form_or_primary_pack_uah",
        "reimbursement_pack_uah",
        "patient_copay_uah",
    ),
}


def detect_section_header(row_values: tuple[Any, ...]) -> str | None:
    """Return the section a row *declares*, or None when the row is not a section header.

    A section can only ever be switched by a sparse, roman-numbered heading row. Data rows
    carry 15-16 populated cells, so they can no longer flip the current section just because
    a trade name happens to contain a word like "комбінований".
    """
    cells = [normalized_text(value) for value in row_values]
    populated = [cell for cell in cells if cell]
    if not populated or len(populated) > SECTION_HEADER_MAX_CELLS:
        return None

    text = " ".join(populated)
    match = SECTION_HEADER_RE.match(text)
    if not match:
        return None

    section = SECTION_ROMAN_NUMERALS.get(normalize_roman_numeral(match.group(1)))
    if section is None:
        return None

    folded = text.casefold()
    keywords = SECTION_HEADER_KEYWORDS[section]
    if not all(keyword in folded for keyword in keywords):
        raise RuntimeError(
            f"Roman-numbered heading {match.group(1)!r} does not match the expected "
            f"{section} section wording: {text[:120]!r}"
        )
    return section


def resolve_section_columns(header_row: tuple[Any, ...], section: str) -> dict[str, int] | None:
    """Map field -> column index for one section's header row.

    Returns None when the row is not a usable header (so the caller can keep scanning), and
    raises when the row looks like a header but the structure changed in a way that would
    silently move values between fields.
    """
    cells = [normalized_cell(value) for value in header_row]
    specs = SECTION_COLUMN_SPECS[section]

    resolved: dict[str, int] = {}
    unmatched: list[str] = []
    ambiguous: list[str] = []

    for field, spec in specs.items():
        exact_values = spec.get("exact", ())
        exact_hits = [index for index, cell in enumerate(cells) if cell and cell in exact_values]
        if len(exact_hits) == 1:
            resolved[field] = exact_hits[0]
            continue

        required = spec.get("all", ())
        hits = [
            index
            for index, cell in enumerate(cells)
            if cell and required and all(token in cell for token in required)
        ]
        if len(hits) == 1:
            resolved[field] = hits[0]
        elif not hits:
            unmatched.append(field)
        else:
            ambiguous.append(f"{field}->{hits}")

    # A row is only a candidate header if it clearly is one; otherwise let the scan continue.
    if len(resolved) < 6:
        return None

    if unmatched or ambiguous:
        problems = []
        if unmatched:
            problems.append(f"unmatched columns: {sorted(unmatched)}")
        if ambiguous:
            problems.append(f"ambiguous columns: {sorted(ambiguous)}")
        raise RuntimeError(
            f"XLSX header for section {section} does not match the expected structure "
            f"({'; '.join(problems)}). Header cells: "
            f"{[cell for cell in cells if cell][:20]}"
        )

    if len(set(resolved.values())) != len(resolved):
        raise RuntimeError(f"XLSX header for section {section} maps two fields to one column: {resolved}")

    return resolved


def is_column_numbering_row(row_values: tuple[Any, ...], columns: dict[str, int]) -> bool:
    """The '1 2 3 4 …' row under each header must not become a data row."""
    inn = normalized_text(row_values[columns["inn_name"]]) if columns["inn_name"] < len(row_values) else ""
    trade = normalized_text(row_values[columns["trade_name"]]) if columns["trade_name"] < len(row_values) else ""
    return bool(re.fullmatch(r"\d+", inn) and re.fullmatch(r"\d+", trade))


def cell_at(row_values: tuple[Any, ...], columns: dict[str, int], field: str) -> Any:
    index = columns.get(field)
    if index is None or index >= len(row_values):
        return None
    return row_values[index]


def build_item(
    row_values: tuple[Any, ...],
    columns: dict[str, int],
    section: str,
    source_row: int,
    ordinal: int,
) -> dict[str, Any]:
    item: dict[str, Any] = {
        "source_row": source_row,
        "section": section,
        "section_uk": SECTIONS[section],
        "ordinal_no": ordinal,
        "inn_name": as_text(cell_at(row_values, columns, "inn_name")),
        "trade_name": as_text(cell_at(row_values, columns, "trade_name")),
        "dosage_form": as_text(cell_at(row_values, columns, "dosage_form")),
        "dosage": as_text(cell_at(row_values, columns, "dosage")),
        "pack_units": as_float(cell_at(row_values, columns, "pack_units")),
        "atc_code": as_text(cell_at(row_values, columns, "atc_code")),
        "manufacturer_country": as_text(cell_at(row_values, columns, "manufacturer_country")),
        "registration_number": as_text(cell_at(row_values, columns, "registration_number")),
        "registration_expiry": as_text(cell_at(row_values, columns, "registration_expiry")),
        "wholesale_price_uah": as_float(cell_at(row_values, columns, "wholesale_price_uah")),
        "retail_price_uah": as_float(cell_at(row_values, columns, "retail_price_uah")),
        "who_daily_dose": None,
        "reimbursement_daily_dose_uah": None,
        "reimbursement_primary_pack_uah": None,
        "reimbursement_release_form_or_primary_pack_uah": None,
        "reimbursement_pack_uah": None,
        "patient_copay_uah": None,
        "copay_type": "",
        "release_form_or_primary_pack": "",
        "icd10_codes": "",
    }

    if section == "insulins":
        item["reimbursement_primary_pack_uah"] = as_float(
            cell_at(row_values, columns, "reimbursement_primary_pack_uah")
        )
        item["patient_copay_uah"] = as_float(cell_at(row_values, columns, "patient_copay_uah"))
        item["copay_type"] = as_text(cell_at(row_values, columns, "copay_type"))
    else:
        copay, copay_label = parse_patient_copay(cell_at(row_values, columns, "patient_copay_uah"))
        item["patient_copay_uah"] = copay
        item["copay_type"] = copay_label
        item["reimbursement_pack_uah"] = as_float(cell_at(row_values, columns, "reimbursement_pack_uah"))
        if section == "medicines":
            item["who_daily_dose"] = as_float(cell_at(row_values, columns, "who_daily_dose"))
            item["reimbursement_daily_dose_uah"] = as_float(
                cell_at(row_values, columns, "reimbursement_daily_dose_uah")
            )
        else:
            item["release_form_or_primary_pack"] = as_text(
                cell_at(row_values, columns, "release_form_or_primary_pack")
            )
            item["reimbursement_release_form_or_primary_pack_uah"] = as_float(
                cell_at(row_values, columns, "reimbursement_release_form_or_primary_pack_uah")
            )

    return item


def validate_required_numeric_fields(rows: list[dict[str, Any]]) -> None:
    for row in rows:
        missing = [field for field in REQUIRED_NUMERIC_FIELDS[row["section"]] if row[field] is None]
        if missing:
            raise RuntimeError(
                f"Missing required numeric values in {row['section']} row {row['source_row']}: {missing}"
            )


def parse_xlsx(path: Path) -> tuple[list[dict[str, Any]], date | None]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if not workbook.worksheets:
        raise RuntimeError("XLSX has no worksheets")

    worksheet = workbook.worksheets[0]
    all_rows = [values for values in worksheet.iter_rows(values_only=True)]

    workbook_register_date: date | None = None
    for values in all_rows[:20]:
        workbook_register_date = parse_strict_register_date_from_text(
            " ".join(normalized_text(value) for value in values)
        )
        if workbook_register_date:
            break
    if workbook_register_date is None:
        raise RuntimeError("XLSX register date is missing from the workbook")

    rows: list[dict[str, Any]] = []
    seen_sections: set[str] = set()
    section: str | None = None
    columns: dict[str, int] | None = None

    for source_row, values in enumerate(all_rows, start=1):
        if not any(normalized_text(value) for value in values):
            continue

        declared = detect_section_header(values)
        if declared is not None:
            if declared in seen_sections:
                raise RuntimeError(f"XLSX declares section {declared} more than once (row {source_row})")
            section = declared
            columns = None
            seen_sections.add(declared)
            continue

        if section is None:
            # Front matter before the first section heading: titles, approval stamps.
            continue

        if columns is None:
            columns = resolve_section_columns(values, section)
            continue

        if is_column_numbering_row(values, columns):
            continue

        ordinal = as_int(cell_at(values, columns, "ordinal_no"))
        if ordinal is None:
            continue

        inn = normalized_text(cell_at(values, columns, "inn_name"))
        trade = normalized_text(cell_at(values, columns, "trade_name"))
        if not inn and not trade:
            continue
        if re.fullmatch(r"\d+", inn) or re.fullmatch(r"\d+", trade):
            raise RuntimeError(
                f"Row {source_row} in section {section} has a numeric name field "
                f"(inn={inn!r}, trade={trade!r}); the column mapping is wrong"
            )

        rows.append(build_item(values, columns, section, source_row, ordinal))

    missing_sections = set(VOLUME_CHECK_SECTIONS) - seen_sections
    if missing_sections:
        raise RuntimeError(f"XLSX missing expected sections: {sorted(missing_sections)}")
    for expected_section in VOLUME_CHECK_SECTIONS:
        if not any(row["section"] == expected_section for row in rows):
            raise RuntimeError(f"XLSX section {expected_section} produced no rows")
    if len(rows) < 100:
        raise RuntimeError(f"XLSX row count is anomalously small: {len(rows)}")
    validate_required_numeric_fields(rows)
    return rows, workbook_register_date


def create_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS metadata (key TEXT PRIMARY KEY, value TEXT);
        CREATE TABLE IF NOT EXISTS reimbursement_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_row INTEGER NOT NULL,
            section TEXT NOT NULL,
            section_uk TEXT NOT NULL,
            ordinal_no INTEGER,
            inn_name TEXT,
            trade_name TEXT,
            dosage_form TEXT,
            dosage TEXT,
            pack_units REAL,
            atc_code TEXT,
            manufacturer_country TEXT,
            registration_number TEXT,
            registration_expiry TEXT,
            wholesale_price_uah REAL,
            retail_price_uah REAL,
            who_daily_dose REAL,
            reimbursement_daily_dose_uah REAL,
            reimbursement_primary_pack_uah REAL,
            reimbursement_release_form_or_primary_pack_uah REAL,
            reimbursement_pack_uah REAL,
            patient_copay_uah REAL,
            copay_type TEXT,
            release_form_or_primary_pack TEXT,
            icd10_codes TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_reimbursement_items_atc ON reimbursement_items(atc_code);
        CREATE INDEX IF NOT EXISTS idx_reimbursement_items_inn ON reimbursement_items(inn_name);
        CREATE INDEX IF NOT EXISTS idx_reimbursement_items_trade ON reimbursement_items(trade_name);
        CREATE INDEX IF NOT EXISTS idx_reimbursement_items_section ON reimbursement_items(section);
        CREATE TABLE IF NOT EXISTS reimbursement_medical_devices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_row INTEGER,
            medical_device_name TEXT,
            medical_device_form TEXT,
            manufacturer TEXT,
            device_description TEXT,
            declaration_info TEXT,
            classifier_code_name TEXT,
            wholesale_price_uah REAL,
            retail_price_uah REAL,
            reimbursement_pack_uah REAL,
            patient_copay_uah REAL
        );
        CREATE TABLE IF NOT EXISTS reference_price_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_row INTEGER NOT NULL,
            section TEXT NOT NULL,
            section_uk TEXT NOT NULL,
            ordinal_no INTEGER,
            inn_name TEXT,
            trade_name TEXT,
            dosage_form TEXT,
            dosage TEXT,
            pack_units REAL,
            atc_code TEXT,
            manufacturer_country TEXT,
            registration_number TEXT,
            registration_expiry TEXT,
            wholesale_price_uah REAL,
            retail_price_uah REAL,
            who_daily_dose REAL,
            reimbursement_daily_dose_uah REAL,
            reimbursement_primary_pack_uah REAL,
            reimbursement_release_form_or_primary_pack_uah REAL,
            reimbursement_pack_uah REAL,
            patient_copay_uah REAL,
            copay_type TEXT,
            release_form_or_primary_pack TEXT,
            icd10_codes TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_reference_price_items_atc ON reference_price_items(atc_code);
        CREATE INDEX IF NOT EXISTS idx_reference_price_items_inn ON reference_price_items(inn_name);
        CREATE INDEX IF NOT EXISTS idx_reference_price_items_trade ON reference_price_items(trade_name);
        CREATE INDEX IF NOT EXISTS idx_reference_price_items_section ON reference_price_items(section);
        """
    )


def table_exists(connection: sqlite3.Connection, table: str) -> bool:
    row = connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table,),
    ).fetchone()
    return row is not None


def table_columns(connection: sqlite3.Connection, table: str) -> set[str]:
    return {row[1] for row in connection.execute(f"PRAGMA table_info({table})").fetchall()}


def prepare_database(
    database_path: Path,
    previous_zip: Path | None,
    *,
    allow_missing_baseline: bool,
) -> bool:
    """Seed the build database from the baseline ZIP. Returns True when a baseline was used.

    Fails closed: a missing or unreadable baseline is only tolerated with an explicit
    override, because silently starting from an empty database drops every table this
    pipeline does not rebuild (reference prices, medical devices, views, curated codes).
    """
    if previous_zip is None or not previous_zip.exists():
        if not allow_missing_baseline:
            raise RuntimeError(
                "Baseline database is required but was not provided. "
                "The previous release ZIP is missing or was not downloaded. "
                "Pass --allow-missing-baseline only when the repository genuinely has no release yet."
            )
        log("Baseline: none (explicitly allowed via --allow-missing-baseline)")
        with sqlite3.connect(database_path) as connection:
            create_schema(connection)
        connection.close()
        return False

    try:
        with zipfile.ZipFile(previous_zip) as archive:
            names = archive.namelist()
            if names.count(SQLITE_NAME) != 1:
                raise RuntimeError(f"Previous ZIP must contain exactly one {SQLITE_NAME}")
            archive.extract(SQLITE_NAME, database_path.parent)
    except zipfile.BadZipFile as error:
        raise RuntimeError(f"Baseline ZIP is corrupted: {previous_zip} ({error})") from error

    extracted = database_path.parent / SQLITE_NAME
    if extracted != database_path:
        extracted.replace(database_path)

    with sqlite3.connect(database_path) as connection:
        integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            raise RuntimeError(f"Baseline SQLite failed integrity_check: {integrity}")
        create_schema(connection)
    connection.close()
    return True


def capture_database_inventory(connection: sqlite3.Connection) -> dict[str, Any]:
    """Row counts per table plus the set of views, taken before the rewrite."""
    tables: dict[str, int] = {}
    for row in connection.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
    ).fetchall():
        table = row[0]
        tables[table] = int(connection.execute(f"SELECT COUNT(*) FROM \"{table}\"").fetchone()[0])

    views = sorted(
        row[0]
        for row in connection.execute("SELECT name FROM sqlite_master WHERE type='view'").fetchall()
    )
    return {"tables": tables, "views": views}


def validate_inventory_preserved(
    previous: dict[str, Any],
    current: dict[str, Any],
) -> None:
    """No table or view that carried data in the baseline may be emptied or dropped."""
    problems: list[str] = []

    for table, previous_count in sorted(previous["tables"].items()):
        current_count = current["tables"].get(table)
        if current_count is None:
            problems.append(f"table {table} was dropped (had {previous_count} rows)")
            continue
        if table in OWNED_TABLES:
            continue
        if previous_count > 0 and current_count == 0:
            problems.append(f"table {table} lost all {previous_count} rows")
        elif current_count < previous_count:
            problems.append(
                f"table {table} lost rows: previous={previous_count}, current={current_count}"
            )

    missing_views = sorted(set(previous["views"]) - set(current["views"]))
    if missing_views:
        problems.append(f"views were dropped: {missing_views}")

    if problems:
        raise RuntimeError("Baseline data would be lost: " + "; ".join(problems))


def item_section_counts(connection: sqlite3.Connection) -> dict[str, int]:
    return {
        row["section"]: int(row["count"])
        for row in connection.execute(
            "SELECT section, COUNT(*) AS count FROM reimbursement_items GROUP BY section"
        ).fetchall()
    }


def metadata_value(connection: sqlite3.Connection, key: str) -> str | None:
    if not table_exists(connection, "metadata"):
        return None
    row = connection.execute("SELECT value FROM metadata WHERE key=?", (key,)).fetchone()
    return normalized_text(row[0]) if row and row[0] is not None else None


@dataclass(frozen=True)
class BaselineProfile:
    """What the baseline database actually contains, and how it was produced."""

    has_items: bool
    section_counts: dict[str, int]
    pipeline_id: str | None
    canonical_source_id: str | None
    icd10_codes_present: int
    icd10_codes_eligible: int

    @property
    def matches_pipeline(self) -> bool:
        return self.pipeline_id == PIPELINE_ID and self.canonical_source_id == CANONICAL_SOURCE_ID

    @property
    def description(self) -> str:
        if not self.has_items:
            return "none (baseline database has no reimbursement_items)"
        if self.matches_pipeline:
            return "previous release from this manual XLSX pipeline"
        return (
            "previous release from a different pipeline/source "
            f"(pipeline_id={self.pipeline_id!r}, canonical_source_id={self.canonical_source_id!r})"
        )


def read_baseline_profile(connection: sqlite3.Connection) -> BaselineProfile:
    """Read the baseline's real counts.

    Deliberately independent of `pipeline_id`: volume and ICD-10 checks must run even when
    the previous release came from another pipeline, otherwise the very first build after a
    source change is the one build with no protection at all.
    """
    if not table_exists(connection, "reimbursement_items"):
        return BaselineProfile(False, {}, None, None, 0, 0)

    section_counts = item_section_counts(connection)
    has_items = sum(section_counts.values()) > 0

    icd10_present = 0
    icd10_eligible = 0
    if "icd10_codes" in table_columns(connection, "reimbursement_items"):
        previous_rows = connection.execute(
            "SELECT section, registration_number, trade_name, dosage_form, dosage, pack_units, "
            "atc_code, icd10_codes FROM reimbursement_items"
        ).fetchall()
        key_counts: dict[tuple[str, ...], int] = {}
        keyed: list[tuple[tuple[str, ...] | None, str]] = []
        for previous in previous_rows:
            key = stable_item_key(previous)
            code = normalized_text(previous["icd10_codes"])
            keyed.append((key, code))
            if key is not None:
                key_counts[key] = key_counts.get(key, 0) + 1
        for key, code in keyed:
            if not code:
                continue
            icd10_present += 1
            if key is not None and key_counts.get(key) == 1:
                icd10_eligible += 1

    return BaselineProfile(
        has_items=has_items,
        section_counts=section_counts,
        pipeline_id=metadata_value(connection, "pipeline_id"),
        canonical_source_id=metadata_value(connection, "canonical_source_id"),
        icd10_codes_present=icd10_present,
        icd10_codes_eligible=icd10_eligible,
    )


def validate_icd10_retention(
    *,
    transferred: int,
    baseline: BaselineProfile,
) -> None:
    """Curated ICD-10 mappings cannot be rebuilt from the workbook, so guard the transfer."""
    if baseline.icd10_codes_eligible == 0:
        return

    minimum = math.ceil(baseline.icd10_codes_eligible * MINIMUM_ICD10_RETENTION_RATIO)
    if transferred < minimum:
        raise RuntimeError(
            "Curated icd10_codes retention dropped below the allowed threshold: "
            f"transferred={transferred}, eligible={baseline.icd10_codes_eligible} "
            f"(of {baseline.icd10_codes_present} present in the baseline), "
            f"minimum={minimum} ({MINIMUM_ICD10_RETENTION_RATIO:.0%} of eligible). "
            "The stable item key stopped matching the workbook; fix the mapping instead of "
            "publishing a database with lost ICD-10 mappings."
        )


def transfer_icd10_codes(connection: sqlite3.Connection, rows: list[dict[str, Any]]) -> int:
    if "icd10_codes" not in table_columns(connection, "reimbursement_items"):
        return 0
    connection.row_factory = sqlite3.Row
    previous_by_key: dict[tuple[str, ...], list[str]] = {}
    for previous in connection.execute(
        "SELECT section, registration_number, trade_name, dosage_form, dosage, pack_units, atc_code, icd10_codes "
        "FROM reimbursement_items"
    ).fetchall():
        key = stable_item_key(previous)
        code = normalized_text(previous["icd10_codes"])
        if key is not None and code:
            previous_by_key.setdefault(key, []).append(code)

    new_key_counts: dict[tuple[str, ...], int] = {}
    new_keys: list[tuple[str, ...] | None] = []
    for row in rows:
        key = stable_item_key(row)
        new_keys.append(key)
        if key is not None:
            new_key_counts[key] = new_key_counts.get(key, 0) + 1

    transferred = 0
    for row, key in zip(rows, new_keys):
        if key is None or new_key_counts[key] != 1:
            continue
        previous_codes = previous_by_key.get(key, [])
        if len(previous_codes) != 1:
            continue
        row["icd10_codes"] = previous_codes[0]
        transferred += 1
    return transferred


def validate_volume_against_previous(
    previous_counts: dict[str, int],
    current_counts: dict[str, int],
) -> None:
    previous_total = sum(previous_counts.get(section, 0) for section in VOLUME_CHECK_SECTIONS)
    current_total = sum(current_counts.get(section, 0) for section in VOLUME_CHECK_SECTIONS)
    if previous_total > 0:
        minimum_total = math.ceil(previous_total * MINIMUM_RELATIVE_COUNT_RATIO)
        if current_total < minimum_total:
            raise RuntimeError(
                "Total reimbursement_items count dropped unexpectedly: "
                f"previous={previous_total}, current={current_total}, minimum={minimum_total}"
            )

    for section in VOLUME_CHECK_SECTIONS:
        previous_count = previous_counts.get(section, 0)
        if previous_count == 0:
            continue
        current_count = current_counts.get(section, 0)
        minimum_count = math.ceil(previous_count * MINIMUM_RELATIVE_COUNT_RATIO)
        if current_count < minimum_count:
            raise RuntimeError(
                f"{section} count dropped unexpectedly: "
                f"previous={previous_count}, current={current_count}, minimum={minimum_count}"
            )


def write_items(connection: sqlite3.Connection, rows: list[dict[str, Any]]) -> None:
    missing_columns = set(ITEM_COLUMNS) - table_columns(connection, "reimbursement_items")
    for column in sorted(missing_columns):
        column_type = ITEM_COLUMN_TYPES.get(column)
        if column_type is None:
            raise RuntimeError(f"No SQL type declared for reimbursement_items column {column!r}")
        connection.execute(f"ALTER TABLE reimbursement_items ADD COLUMN {column} {column_type}")
    connection.execute("DELETE FROM reimbursement_items")
    insert_columns = ", ".join(ITEM_COLUMNS)
    insert_values = ", ".join(f":{column}" for column in ITEM_COLUMNS)
    connection.executemany(
        f"INSERT INTO reimbursement_items ({insert_columns}) VALUES ({insert_values})",
        rows,
    )


# Metadata written by an earlier pipeline is only kept when it describes data this build still
# carries over. `reference_prices_*` documents `reference_price_items`, which survives untouched;
# everything else (source_pdf, source_json, preliminary_*, per-run counters of the old PDF
# pipeline) would otherwise make the database lie about its own provenance.
PERSISTENT_METADATA_PREFIXES = ("reference_prices_",)


def write_metadata(connection: sqlite3.Connection, source: SourceCandidate, rows: list[dict[str, Any]]) -> list[str]:
    section_counts = item_section_counts(connection)
    devices_count = (
        connection.execute("SELECT COUNT(*) FROM reimbursement_medical_devices").fetchone()[0]
        if table_exists(connection, "reimbursement_medical_devices")
        else 0
    )
    metadata = {
        "pipeline_id": PIPELINE_ID,
        "canonical_source_id": CANONICAL_SOURCE_ID,
        "document_title": "Перелік лікарських засобів, які підлягають реімбурсації",
        "register_date": source.register_date.isoformat(),
        "source_file": source.name,
        "source_url": source.url,
        "source_resource_id": source.resource_id,
        "source_description": source.description,
        "items_count": str(len(rows)),
        "devices_count": str(devices_count),
        "record_count_by_section": json.dumps(section_counts, ensure_ascii=False, sort_keys=True),
        "sections": json.dumps(SECTIONS, ensure_ascii=False, sort_keys=True),
        "updated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }
    connection.executemany(
        """
        INSERT INTO metadata(key, value)
        VALUES(?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """,
        metadata.items(),
    )

    existing = {
        row[0]
        for row in connection.execute("SELECT key FROM metadata").fetchall()
        if row and row[0] is not None
    }
    stale = sorted(
        key
        for key in existing
        if key not in metadata and not key.startswith(PERSISTENT_METADATA_PREFIXES)
    )
    if stale:
        connection.executemany(
            "DELETE FROM metadata WHERE key = ?",
            ((key,) for key in stale),
        )
    return stale


def validate_sqlite(database_path: Path) -> dict[str, int]:
    with sqlite3.connect(database_path) as connection:
        connection.row_factory = sqlite3.Row
        integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            raise RuntimeError(f"SQLite integrity_check failed: {integrity}")
        for table in ("metadata", "reimbursement_items"):
            if not table_exists(connection, table):
                raise RuntimeError(f"Missing required table: {table}")
        columns = table_columns(connection, "reimbursement_items")
        missing = REQUIRED_ITEM_COLUMNS - columns
        if missing:
            raise RuntimeError(f"Missing reimbursement_items columns: {sorted(missing)}")
        section_counts = {
            row["section"]: int(row["count"])
            for row in connection.execute(
                "SELECT section, COUNT(*) AS count FROM reimbursement_items GROUP BY section"
            ).fetchall()
        }
        items_count = sum(section_counts.values())
        if items_count < 100:
            raise RuntimeError(f"reimbursement_items count is anomalously small: {items_count}")
        if section_counts.get("medicines", 0) < 50:
            raise RuntimeError("medicines section is anomalously small")
        if section_counts.get("insulins", 0) < 10:
            raise RuntimeError("insulins section is anomalously small")
        devices_count = 0
        if table_exists(connection, "reimbursement_medical_devices"):
            devices_count = connection.execute("SELECT COUNT(*) FROM reimbursement_medical_devices").fetchone()[0]
        if table_exists(connection, "reference_price_items"):
            _ = connection.execute("SELECT COUNT(*) FROM reference_price_items").fetchone()[0]
        result = {
            "items": items_count,
            "medicines": section_counts.get("medicines", 0),
            "insulins": section_counts.get("insulins", 0),
            "combined_medicines": section_counts.get("combined_medicines", 0),
            "medical_devices": devices_count,
        }
    connection.close()
    return result


def make_zip(database_path: Path, zip_path: Path) -> str:
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
        archive.write(database_path, SQLITE_NAME)
    with zipfile.ZipFile(zip_path) as archive:
        if archive.namelist() != [SQLITE_NAME]:
            raise RuntimeError(f"ZIP must contain only {SQLITE_NAME}")
    digest = hashlib.sha256()
    with zip_path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_manifest(path: Path, *, database_version: int, source_date: date, repo: str, sha256: str) -> None:
    tag = f"reimbursement-v{database_version}"
    manifest = {
        "databaseVersion": database_version,
        "schemaVersion": SCHEMA_VERSION,
        "updatedAt": source_date.isoformat(),
        "minimumAppVersion": MINIMUM_APP_VERSION,
        "downloadURL": f"https://github.com/{repo}/releases/download/{tag}/{ZIP_NAME}",
        "sha256": sha256,
    }
    path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_summary(
    path: Path,
    stats: BuildStats,
    reason: str,
    *,
    inventory: dict[str, Any] | None = None,
    removed_metadata_keys: list[str] | None = None,
) -> None:
    path.write_text(
        json.dumps(
            {
                "publish": False,
                "source_version": stats.source_version,
                "source_file": stats.source_file,
                "rows_processed": stats.rows_processed,
                "medicines_count": stats.medicines_count,
                "insulins_count": stats.insulins_count,
                "combined_medicines_count": stats.combined_medicines_count,
                "medical_devices_count": stats.medical_devices_count,
                "database_version": stats.database_version,
                "sha256": stats.sha256,
                "icd10_codes_transferred": stats.icd10_codes_transferred,
                "icd10_codes_previous": stats.icd10_codes_previous,
                "icd10_codes_eligible": stats.icd10_codes_eligible,
                "icd10_retention_threshold": MINIMUM_ICD10_RETENTION_RATIO,
                "baseline": stats.baseline,
                "expected_release_tag": stats.expected_release_tag,
                "release_created": stats.release_created,
                "carried_over_tables": {
                    table: count
                    for table, count in sorted((inventory or {}).get("tables", {}).items())
                    if table not in OWNED_TABLES
                },
                "carried_over_views": (inventory or {}).get("views", []),
                "removed_stale_metadata_keys": removed_metadata_keys or [],
                "reason": reason,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build DocSPACE reimbursement release assets.")
    parser.add_argument("--work-dir", type=Path, required=True)
    parser.add_argument("--repo", required=True, help="GitHub repository, for example owner/name")
    parser.add_argument("--input-dir", type=Path, required=True)
    parser.add_argument("--previous-manifest", type=Path)
    parser.add_argument("--previous-zip", type=Path)
    parser.add_argument(
        "--minimum-database-version",
        type=int,
        default=0,
        help="Highest database version that already exists as a release. The build refuses to "
             "produce a version that is not strictly greater, so an unreachable baseline can "
             "never re-issue an already published version.",
    )
    parser.add_argument(
        "--allow-missing-baseline",
        action="store_true",
        help="Build without a previous release. Only for a repository with no releases at all.",
    )
    parser.add_argument(
        "--allow-pipeline-transition",
        action="store_true",
        help="Accept a baseline produced by a different pipeline/source. Volume and ICD-10 checks "
             "still run against the baseline's real counts.",
    )
    parser.add_argument(
        "--allow-same-source-date",
        action="store_true",
        help="Rebuild from a workbook whose register date is not newer than the published one.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    args.work_dir.mkdir(parents=True, exist_ok=True)
    summary_path = args.work_dir / "build_summary.json"

    # B5: versioning fails closed. Without a baseline manifest the next version is unknowable,
    # so guessing 1 (and colliding with already published v1..vN) is not an option.
    manifest = latest_published_manifest(args.previous_manifest)
    if manifest is None and not args.allow_missing_baseline:
        raise RuntimeError(
            "Baseline manifest is required but was not provided. Refusing to guess "
            "databaseVersion=1, which would collide with already published releases. "
            "Pass --allow-missing-baseline only when the repository has no release yet."
        )
    previous_version = int(manifest["databaseVersion"]) if manifest else 0
    previous_source_date = parse_manifest_date(manifest.get("updatedAt") if manifest else None)

    xlsx_path = find_input_workbook(args.input_dir)
    rows, workbook_register_date = parse_xlsx(xlsx_path)
    if workbook_register_date is None:
        raise RuntimeError("XLSX register date is missing from the workbook")
    source = SourceCandidate(
        name=xlsx_path.name,
        url="",
        resource_id="",
        register_date=workbook_register_date,
        description="Manual XLSX input from input/reimbursement",
    )
    log(f"Source version: {source.register_date.isoformat()} ({source.name})")

    database_path = args.work_dir / SQLITE_NAME
    used_baseline = prepare_database(
        database_path,
        args.previous_zip,
        allow_missing_baseline=args.allow_missing_baseline,
    )

    with sqlite3.connect(database_path) as connection:
        connection.row_factory = sqlite3.Row
        baseline = read_baseline_profile(connection)
        baseline_inventory = capture_database_inventory(connection)
    connection.close()
    log(f"Volume baseline: {baseline.description}")
    if used_baseline and baseline.has_items and not baseline.matches_pipeline:
        # B3/B4: a cross-pipeline baseline is exactly the case that used to disable every
        # relative check. It now requires an explicit decision, and the checks still run.
        if not args.allow_pipeline_transition:
            raise RuntimeError(
                "Baseline was produced by a different pipeline/source: "
                f"pipeline_id={baseline.pipeline_id!r} (expected {PIPELINE_ID!r}), "
                f"canonical_source_id={baseline.canonical_source_id!r} (expected {CANONICAL_SOURCE_ID!r}). "
                "Volume and ICD-10 checks will run against its real counts, but row counts and the "
                "register date are not directly comparable across sources. "
                "Re-run with --allow-pipeline-transition once you have confirmed the drop is expected."
            )
        log("Pipeline transition: explicitly allowed via --allow-pipeline-transition")

    # B4: the date gate only means something when both sides come from the same source.
    date_comparable = baseline.matches_pipeline or not baseline.has_items
    if previous_source_date and source.register_date <= previous_source_date:
        if not date_comparable and args.allow_pipeline_transition:
            log(
                f"Register date {source.register_date.isoformat()} is not newer than published "
                f"{previous_source_date.isoformat()}, but the baseline came from a different source "
                "so the dates are not comparable; continuing."
            )
        elif args.allow_same_source_date:
            log(
                f"Register date {source.register_date.isoformat()} is not newer than published "
                f"{previous_source_date.isoformat()}; continuing via --allow-same-source-date."
            )
        else:
            database_version = previous_version
            stats = BuildStats(
                source_version=source.register_date.isoformat(),
                source_file=source.name,
                rows_processed=len(rows),
                medicines_count=0,
                insulins_count=0,
                combined_medicines_count=0,
                medical_devices_count=0,
                database_version=database_version,
                sha256="",
                icd10_codes_transferred=0,
                icd10_codes_previous=baseline.icd10_codes_present,
                icd10_codes_eligible=baseline.icd10_codes_eligible,
                baseline=baseline.description,
                expected_release_tag="",
                release_created=False,
            )
            reason = (
                f"No changes: manual XLSX source {source.register_date.isoformat()} is not newer "
                f"than published {previous_source_date.isoformat()}. "
                "Re-run with --allow-same-source-date to rebuild the same date."
            )
            # The baseline copy was extracted into the work dir to read its profile. Remove it,
            # otherwise the artifact upload would ship the *previous* database as if it were a
            # build output of this run.
            database_path.unlink(missing_ok=True)
            log(reason)
            write_summary(summary_path, stats, reason)
            return 0

    database_version = previous_version + 1
    if database_version <= args.minimum_database_version:
        raise RuntimeError(
            f"Refusing to build databaseVersion={database_version}: release "
            f"reimbursement-v{args.minimum_database_version} already exists. "
            "The baseline manifest is stale or was not the latest release."
        )

    with sqlite3.connect(database_path) as connection:
        connection.row_factory = sqlite3.Row
        connection.execute("BEGIN")
        transferred_icd10_codes = transfer_icd10_codes(connection, rows)
        write_items(connection, rows)
        removed_metadata_keys = write_metadata(connection, source, rows)
        connection.commit()
        connection.execute("VACUUM")
    connection.close()
    counts = validate_sqlite(database_path)

    with sqlite3.connect(database_path) as connection:
        current_inventory = capture_database_inventory(connection)
    connection.close()
    # B2: nothing the baseline carried may be emptied or dropped by the rewrite.
    if used_baseline:
        validate_inventory_preserved(baseline_inventory, current_inventory)

    # B3: always compare against the baseline's real counts when it had any items.
    if baseline.has_items:
        validate_volume_against_previous(baseline.section_counts, counts)
    else:
        log("Volume check: skipped, baseline has no reimbursement_items to compare against")

    # B6: curated ICD-10 mappings must survive.
    validate_icd10_retention(transferred=transferred_icd10_codes, baseline=baseline)

    zip_path = args.work_dir / ZIP_NAME
    sha256 = make_zip(database_path, zip_path)
    expected_release_tag = f"reimbursement-v{database_version}"
    write_manifest(
        args.work_dir / MANIFEST_NAME,
        database_version=database_version,
        source_date=source.register_date,
        repo=args.repo,
        sha256=sha256,
    )

    stats = BuildStats(
        source_version=source.register_date.isoformat(),
        source_file=source.name,
        rows_processed=len(rows),
        medicines_count=counts["medicines"],
        insulins_count=counts["insulins"],
        combined_medicines_count=counts["combined_medicines"],
        medical_devices_count=counts["medical_devices"],
        database_version=database_version,
        sha256=sha256,
        icd10_codes_transferred=transferred_icd10_codes,
        icd10_codes_previous=baseline.icd10_codes_present,
        icd10_codes_eligible=baseline.icd10_codes_eligible,
        baseline=baseline.description,
        expected_release_tag=expected_release_tag,
        release_created=False,
    )
    log(f"Rows processed: {stats.rows_processed}")
    log(f"Medicines count: {stats.medicines_count}")
    log(f"Insulins count: {stats.insulins_count}")
    log(f"Combined medicines count: {stats.combined_medicines_count}")
    log(f"Medical devices count: {stats.medical_devices_count}")
    log(
        f"ICD10 codes transferred: {transferred_icd10_codes} "
        f"(eligible {baseline.icd10_codes_eligible} of {baseline.icd10_codes_present} in baseline)"
    )
    log(f"Reference price items: {current_inventory['tables'].get('reference_price_items', 0)}")
    if removed_metadata_keys:
        log(f"Removed stale provenance metadata keys: {removed_metadata_keys}")
    log(f"Database version: {stats.database_version}")
    log(f"Expected release tag: {expected_release_tag}")
    log(f"SHA256: {stats.sha256}")
    log("Release creation: disabled; build artifacts are ready")
    write_summary(
        summary_path,
        stats,
        "Build artifacts ready; GitHub Release creation is disabled",
        inventory=current_inventory,
        removed_metadata_keys=removed_metadata_keys,
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (RuntimeError, zipfile.BadZipFile, sqlite3.Error) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise SystemExit(1)
